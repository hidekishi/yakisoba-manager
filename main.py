import openpyxl as xl
import os

os.system('cls')
nome_tabela = input("Nome da tabela>> ")
plan = xl.load_workbook(f'{nome_tabela}.xlsx')

def print_line(num_ficha, sheet):
    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
        if row[0].value == num_ficha:
            row_values = [cell.value for cell in row]
            print(row_values)
            break

def get_num_ficha(): # Retorna o numero e o nome da planilha
    num_ficha = int(input("Número da ficha>> "))
    while(num_ficha < 1 or num_ficha > 900):
        os.system('cls')
        num_ficha = int(input("Número de ficha inexistente. Tente de novo>> "))
    if num_ficha <= 200:
        sheet_name = 'Sheet1'
    elif num_ficha <= 400:
        sheet_name = '201-400'
    elif num_ficha <= 600:
        sheet_name = '401-600'
    elif num_ficha <= 800:
        sheet_name = '601-800'
    else:
        sheet_name = '801-900'
    return num_ficha, sheet_name

def ent_mod_ficha(): # Entrada e modificacao de dados da ficha
    while(True):
        os.system('cls')
        num_ficha, sheet_name = get_num_ficha()
        sheet = plan[sheet_name]
        print(num_ficha, sheet)
        break

def info_ficha(): # Exibicao de dados da ficha 
    while(True):
        os.system('cls')
        num_ficha, sheet_name = get_num_ficha()
        sheet = plan[sheet_name]
        print_line(num_ficha, sheet)
        break


while(True):
    #os.system('cls')
    cmd = int(input('Selecione a tarefa a ser realizada:\n(1) Entrada ou modificação de ficha\n(2) Informações de ficha\n(3) Sair\nTarefa>> '))
    if cmd == 1:
        ent_mod_ficha()
    elif cmd == 2:
        info_ficha()
    elif cmd == 3:
        break
    else:
        input('Tarefa inválida. Pressione <Enter> e tente novamente!')

#